# -*- coding: utf-8 -*-
"""
Created on Thu Jul 27 16:03:13 2023

@author: kkeshi
"""

#%% Module imports
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
import os
import category_encoders as ce
import optuna
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.model_selection import train_test_split, StratifiedKFold, KFold, cross_validate, GridSearchCV
from sklearn.metrics import make_scorer, recall_score, precision_score, f1_score, fbeta_score, accuracy_score, roc_auc_score, confusion_matrix
from sklearn.metrics import roc_curve, precision_recall_curve
from sklearn.metrics import mean_absolute_percentage_error, r2_score, mean_squared_error
from sklearn.linear_model import LogisticRegression, LinearRegression
from sklearn.calibration import calibration_curve
from sklearn.feature_selection import SelectKBest, mutual_info_classif, mutual_info_regression
from imblearn.over_sampling import SVMSMOTE, RandomOverSampler
from imblearn.metrics import geometric_mean_score
from xgboost import XGBClassifier, XGBRegressor
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from plotly.offline import plot
import re
from openpyxl.utils.dataframe import dataframe_to_rows


#%% User Inputs
dataset_filename = 'world-data-2023.csv' # File name including extension 
use_cols = 'b, d:h, J:l, n:t, v:x, z:ai' # State desired column indices. Format is 'C:F, I, J'. Defaults to using all
target = 'Life expectancy' # State the variable you want to model for
modeltype = 'R' # State the type of problem this is - 'C' for Classification 'R' for Regression
force_features = [] # Based on your own knowledge enter the column headers in Excel file of features you wish to bypass feature selection. Example: 'age'
QuickSearch = '' # Type 'Y' for Quick Search - a significantly faster but unoptimised model. Defaults to optimised model.
focus_metric = '' # The metric you wish to optimise for, for classification options are: accuracy, precision, recall, f1, f1/2, f2, roc_auc, g_mean. Defaults to f1
                  # For regression options are: mean_squared_error, mean_absolute_percentage_error, r2. Defaults to r2
low_memory = ''   # For less memory consumption set to 'Y', limits row count to 10000



#%% Load dataset
current_location = os.path.dirname(os.path.abspath(__file__))
dataset_filepath = current_location + "\\" + dataset_filename

if low_memory == 'Y':
    nrows=10000
else:
    nrows = None
def parse_excel_columns(input_string):
    values = []
    ranges = input_string.split(',')

    for col_range in ranges:
        col_range = col_range.strip()

        if ':' in col_range:
            start_col, end_col = col_range.split(':')
            start_col = start_col.strip()
            end_col = end_col.strip()

            start_value = column_value(start_col)
            end_value = column_value(end_col)

            for value in range(start_value, end_value + 1):
                values.append(value)
        else:
            values.append(column_value(col_range))

    return values

def column_label(col_index):
    col_label = ''
    while col_index >= 0:
        col_label = chr(col_index % 26 + ord('A')) + col_label
        col_index = col_index // 26 - 1
    return col_label

def column_value(col_name):
    base = 26
    result = -1
    for idx, char in enumerate(col_name):
        order = len(col_name) - idx - 1
        result += base ** order * (ord(char.upper()) - ord('A') + 1)
    return result
        
try:    
    if use_cols is None or use_cols == '':
        df = pd.read_csv(dataset_filepath)
    else:
        columns_to_use = parse_excel_columns(use_cols)
        df = pd.read_csv(dataset_filepath, usecols=columns_to_use, nrows=nrows)
except UnicodeDecodeError:
    if use_cols is None or use_cols == '':
        df = pd.read_excel(dataset_filepath, nrows=nrows)
    else:
        df = pd.read_excel(dataset_filepath, usecols=use_cols, nrows=nrows)
#%% Preprocessing
#Remove certain symbols so the column entry can be used
def remove_symbols(entry):
    if isinstance(entry, str):  # Check if entry is a string
        cleaned_entry = re.sub(r'[$€£¥₹%,]', '', entry)
        try:
            return float(cleaned_entry)
        except ValueError:  # If not a float (e.g., if the entry is empty), return as it is
            return cleaned_entry
    return entry  # Return entry as it is if not a string

df = df.applymap(remove_symbols)
df = df.apply(pd.to_numeric, errors='ignore')

# Drop empty columns
df = df.dropna(axis=1, how='all')

# Drop rows without entry for target
df = df.dropna(subset=target, how='all')

#Dealing with Categorical variables
#This process makes the more common class 0 and rare class 1 - important for precision and recall
if df[target].dtype == 'object' or df[target].dtype == 'bool':
    if sum(df[target] == pd.unique(df[target])[0]) < 0.5 * len(df):
        df.loc[df[target] == pd.unique(df[target])[0], target] = 1
        df.loc[df[target] == pd.unique(df[target])[1], target] = 0
    else:
        df.loc[df[target] == pd.unique(df[target])[1], target] = 1
        df.loc[df[target] == pd.unique(df[target])[0], target] = 0
    df[target] = pd.to_numeric(df[target])

cat_features = df.select_dtypes(include='object').columns.tolist()
bool_features = df.select_dtypes(include='bool').columns.tolist()

# Feature encoding for correlation matrices
df_encoded = df
if len(cat_features) == 0:
    pass
else:
    catboost_encoder = ce.CatBoostEncoder(cols=cat_features)
    df_encoded = catboost_encoder.fit_transform(df, df[target])

if len(bool_features) == 0:
    pass
else:
    label_encoder = LabelEncoder()
    for boolean in bool_features:
        df_encoded[boolean] = label_encoder.fit_transform(df_encoded[boolean])

#%% Correlation Analysis
corr = df_encoded.corr(numeric_only=True)
#Export correlation matrix 
#Reducing decimal places
def convert(entry):
    if isinstance(entry, float):
        return round(entry, 2)

corr = corr.applymap(convert)
filename = dataset_filename.split('.')[0]
corr.to_excel(current_location + '\\' f'{filename} Modelling.xlsx', sheet_name='Correlation Matrix', index=True)
workbook = openpyxl.load_workbook(current_location + '\\' f'{filename} Modelling.xlsx')
worksheet = workbook['Correlation Matrix']
#Introduce colour for 
worksheet.conditional_formatting.add(f'B2:{column_label(len(corr))}{len(corr) + 1}',
                                     ColorScaleRule(start_type='num', start_value=-1, start_color='0045FE',
                                                    mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                    end_type='num', end_value=1, end_color='FF0000'))

#Addressing multicollinearity
featureframe = df_encoded.drop(columns=target)
corr_feature = featureframe.corr(numeric_only=True)
threshold = 0.75 # If the correlation between two non-target features is above this threshold, they need to be addressed
remove_features = []
for feature in corr_feature.columns:
    if feature not in remove_features:
        for feature2 in corr_feature.columns:
            if feature != feature2 and abs(corr_feature[feature][feature2]) > threshold:
                if feature in force_features and feature2 in force_features:
                    continue
                elif feature2 not in force_features:
                    remove_features.append(feature2)
                    featureframe.drop(columns=feature2, inplace=True)
        #Update the feature frame correlation matrix
        corr_feature = featureframe.corr(numeric_only=True)
        
#Feature frame after dealing with multicollinearity
corr_updated = corr_feature

selected_features = [feature for feature in corr_feature.columns]
selected_features.append(target)
df_selected = df[selected_features]         
df_selected = df_selected.dropna()   
#%% Initialisation
datapoint_count = len(df_selected[target]) # Number of datapoints
variable_count = len(df_selected.columns) # Number of variables 
X = df_selected.drop(columns=target)
Y = df_selected[target]
#Train-test splitting
if modeltype == 'C':
    X_train, X_test, Y_train, Y_test = train_test_split(X, Y, stratify=Y, test_size=0.3)
    weights = Y.value_counts(normalize=True)
else:
    X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.3)
    
k = None

if 10 * variable_count > datapoint_count: # If datapoints:features is less than 10:1, we reduce the feature count. 
    if datapoint_count < 1000:
        k = int(datapoint_count / 10) # Upper threshold 99
    elif 10000 > datapoint_count >= 1000:
        k = int(max(((np.log2(datapoint_count))**2), datapoint_count / 50)) # Lower threshold: 99, Upper: 200
    elif 100000 > datapoint_count >= 10000:
        k = int(max((8/7*(np.log2(datapoint_count))**2), datapoint_count / 200)) # Lower threshold: 201, Upper: 500
    else:
        k = 500 # Cap at 500 variables

#Feature selection using Mutual Information
if k is not None:
    if modeltype == 'C':
        kbest = SelectKBest(score_func=mutual_info_classif, k=k)
    elif modeltype == 'R':
        kbest = SelectKBest(score_func=mutual_info_regression, k=k)
    else:
        raise ValueError('Model type not specified - please input "C" for classification tasks or "R" for regression tasks')
    
    #Scaling and encoding on the entire dataset for mutual information
    X_train_encoded = X_train
    cat_features = X_train.select_dtypes(include='object').columns.tolist()
    bool_features = X_train.select_dtypes(include='bool').columns.tolist()
    if len(cat_features) == 0:
        pass
    else:
        catboost_encoder = ce.CatBoostEncoder(cols=cat_features)
        X_train_encoded = catboost_encoder.fit_transform(X_train, Y_train) 
    if len(bool_features) == 0:
        pass
    else:
        for boolean in bool_features:
            X_train_encoded[boolean] = label_encoder.fit_transform(X_train_encoded[boolean])
        
    scaler = StandardScaler() 
    X_train_scaled = scaler.fit_transform(X_train_encoded)
    X_scaled_kbest_selected = kbest.fit_transform(X_train_scaled, Y_train)
    kbest_feature_indices = kbest.get_support()
    kbest_names = X.columns[kbest_feature_indices]
    force_features = pd.Index(force_features)
    for feature in force_features:
        if feature not in kbest_names:
            kbest_names = kbest_names.union([feature])
    X_train, X_test = X_train[kbest_names], X_test[kbest_names] 
else:
    pass

#Cross-Validation Methods
skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
kf = KFold(n_splits=5, shuffle=True, random_state=42)

X_train = pd.get_dummies(X_train)
X_test = pd.get_dummies(X_test)

# Handling missing classes for features
missing_columns_train = list(set(X_test.columns) - set(X_train.columns))
missing_columns_test = list(set(X_train.columns) - set(X_test.columns))

missing_columns_train_df = pd.DataFrame(0, index=X_train.index, columns=missing_columns_train)
X_train = pd.concat([X_train, missing_columns_train_df], axis=1)

missing_columns_test_df = pd.DataFrame(0, index=X_test.index, columns=missing_columns_test)
X_test = pd.concat([X_test, missing_columns_test_df], axis=1)
X_test = X_test[X_train.columns]

def geometric_mean(scores):
    product = 1.0
    for score in scores:
        product *= score
    return product ** (1.0 / len(scores))

models = ['linear', 'xgboost']
#%% Classification
#Following actions depend on whether this is a classification task or regression task
if modeltype == 'C':
    scoring = { #Performance metrics used in fitting
        'accuracy': make_scorer(accuracy_score),
        'precision': make_scorer(precision_score),
        'recall': make_scorer(recall_score),
        'f1': make_scorer(f1_score),
        'f1/2': make_scorer(fbeta_score, beta=0.5),
        'f2': make_scorer(fbeta_score, beta=2),
        'roc_auc': make_scorer(roc_auc_score),
        'g_mean': make_scorer(geometric_mean_score)
    }

    if min(weights) * 10 < max(weights): # Only want to consider class imbalance if this imbalance is severe
        try:
            resampler = SVMSMOTE(random_state=42)
            X_train_resampled, Y_train_resampled = resampler.fit_resample(X_train, Y_train)
        except ValueError:
            resampler = RandomOverSampler(random_state=42)
            X_train_resampled, Y_train_resampled = resampler.fit_resample(X_train, Y_train)
    else:
        X_train_resampled, Y_train_resampled = X_train, Y_train


    def Classification(models, QuickSearch):
        scoring = { #Performance metrics used in fitting
            'accuracy': make_scorer(accuracy_score),
            'precision': make_scorer(precision_score),
            'recall': make_scorer(recall_score),
            'f1': make_scorer(f1_score),
            'f1/2': make_scorer(fbeta_score, beta=0.5),
            'f2': make_scorer(fbeta_score, beta=2),
            'roc_auc': make_scorer(roc_auc_score),
            'g_mean': make_scorer(geometric_mean_score)
        }
        
        # Feature scaling    
        scaler = StandardScaler()
        X_train_fit = scaler.fit_transform(X_train_resampled)
        X_test_fit = scaler.transform(X_test)
        results = []

        for model in models:
            if model == 'linear':
                alg = LogisticRegression(dual=False, max_iter=100000, random_state=42)
                if QuickSearch != 'Y':
                    # Optimising the Logistic Regression Model
                    def evaluate_log_reg(trial):
                        log_reg_space = { # Hyperparameter space for Logistic Regression
                            'C': trial.suggest_float('C', 0.001, 10.0, log=True),  # Penalty parameter C with a log-uniform distribution
                            'penalty': trial.suggest_categorical('penalty', ['l1', 'l2', 'elasticnet']),  # Regularization penalty ('l1' or 'l2')
                            'tol':  trial.suggest_float('tol', 0.0001, 0.1, log=True),  # Tolerance for stopping criteria
                            'fit_intercept': trial.suggest_categorical('fit_intercept', [True, False]),  # Whether to calculate the intercept for the model
                            'class_weight': trial.suggest_categorical('class_weight', ['balanced', None]),
                        }
                        if len(X_train_fit) < 10000:
                            log_reg_space['solver'] = trial.suggest_categorical('solver', ['liblinear', 'lbfgs', 'newton-cg'])
                        else:
                            log_reg_space['solver'] = trial.suggest_categorical('solver', ['newton-cholesky', 'saga'])
                        
                        if log_reg_space['penalty'] == 'elasticnet':
                            log_reg_space['l1_ratio'] = trial.suggest_float('l1_ratio', 0.01, 1.0)
                        
                        log_reg_model = LogisticRegression(**log_reg_space, dual=False, max_iter=100000, random_state=42)
                        
                        try:
                            log_reg_model.fit(X_train_fit, Y_train_resampled)
                        except ValueError:
                            raise optuna.TrialPruned()
                        
                        result = cross_validate(log_reg_model, X_train_fit, Y_train_resampled, cv=skf, scoring=scoring, n_jobs=-1)

                        if focus_metric != '':
                            score = geometric_mean(result[f'test_{focus_metric}'])
                            return score
                        else: # Make suggestion based on class distributions if there is 
                            f1 = geometric_mean(result['test_f1'])
                            return f1
                    study_log_reg = optuna.create_study(direction='maximize')
                    study_log_reg.optimize(evaluate_log_reg, n_trials=200)
                    best_log_reg_params = study_log_reg.best_params
                    alg = LogisticRegression(**best_log_reg_params, dual=False, max_iter=100000, random_state=42)
                alg.fit(X_train_fit, Y_train_resampled)
            else:
                alg = XGBClassifier(n_estimators=1000, learning_rate=0.01, subsample=0.5, n_jobs=-1, random_state=42)
                if QuickSearch != 'Y':
                    # Optimising XGBoost
                    def evaluate_xgb(trial):
                        xgb_space = { # Hyperparameter space for XGBoost
                            'n_estimators': trial.suggest_int('n_estimators', 500, 1500),  # Number of boosting rounds
                            'max_depth': trial.suggest_int('max_depth', 1, 3),  # Maximum depth of a tree
                            'learning_rate': trial.suggest_float('learning_rate', 0.001, 0.01, log=True),  # Step size shrinkage
                            'gamma': trial.suggest_float('gamma', 0.0, 1.0),  # Minimum loss reduction required to make a further partition
                            'colsample_bytree': trial.suggest_float('colsample_bytree', 0.1, 1.0),  # Fraction of features used for fitting the trees
                            'reg_alpha': trial.suggest_float('reg_alpha', 0.0, 100.0),  # L1 regularization term on weights
                            'reg_lambda': trial.suggest_float('reg_lambda', 0.0, 100.0),  # L2 regularization term on weights
                        }
                        xgb_model = XGBClassifier(**xgb_space, subsample=0.5, n_jobs=-1, random_state=42)
                        
                        try:
                            xgb_model.fit(X_train_fit, Y_train_resampled)
                        except ValueError:
                            raise optuna.TrialPruned()
                        
                        result = cross_validate(xgb_model, X_train_fit, Y_train_resampled, cv=skf, scoring=scoring, n_jobs=-1)

                        if focus_metric != '':
                            score = geometric_mean(result[f'test_{focus_metric}'])
                            return score
                        else: # Make suggestion based on class distributions if there is 
                            f1 = geometric_mean(result['test_f1'])
                            return f1
                    study_xgb = optuna.create_study(direction='maximize')
                    study_xgb.optimize(evaluate_xgb, n_trials=200)
                    best_xgb_params = study_xgb.best_params
                    alg = XGBClassifier(**best_xgb_params, random_state=42)
                alg.fit(X_train_fit, Y_train_resampled)
            results.append(alg)
            
            pred = alg.predict(X_test_fit)
            results.append(pred)
            probabilities = alg.predict_proba(X_test_fit)[:, 1]
            roc = roc_curve(Y_test, probabilities)
            pr = precision_recall_curve(Y_test, probabilities)
            cal = calibration_curve(Y_test, probabilities, n_bins=10)
            results.append(roc)
            results.append(pr)
            results.append(cal)
            
        return results
    
    log_reg, log_reg_pred, log_reg_roc, log_reg_pr, log_reg_cal, xgb, xgb_pred, xgb_roc, xgb_pr, xgb_cal = Classification(models, QuickSearch)
    
    # Unpacking Logistic Regression outputs
    fpr_log_reg, tpr_log_reg, thresholds_log_reg = log_reg_roc
    precision_log_reg, recall_log_reg, thresholds_pr_log_reg = log_reg_pr
    true_log_reg_probs, predicted_log_reg_probs = log_reg_cal
    
    # Unpacking XGBoost outputs
    fpr_xgb, tpr_xgb, thresholds_xgb = xgb_roc
    precision_xgb, recall_xgb, thresholds_pr_xgb = xgb_pr
    true_xgb_probs, predicted_xgb_probs = xgb_cal
    
    #Export Data to Excel File
    #ROC Curve Data
    xgb_roc_df = pd.DataFrame({'Threshold': thresholds_xgb, 'False Positive Rate': fpr_xgb, 'True Positive Rate': tpr_xgb})
    xgb_roc_rows = list(dataframe_to_rows(xgb_roc_df, index=False))
    log_reg_roc_df = pd.DataFrame({'Threshold': thresholds_log_reg, 'False Positive Rate': thresholds_log_reg, 'True Positive Rate': thresholds_log_reg})
    log_reg_roc_rows = list(dataframe_to_rows(log_reg_roc_df, index=False))
    worksheet = workbook.create_sheet(title=f'ROC Curve Data - {target}')
    title_row_xgb, title_row_log_reg = [' ', 'XGBoost', ' '], [' ', 'Logistic Regression', ' ']
    
    worksheet.append(title_row_xgb)
    for row in xgb_roc_rows:
        worksheet.append(row)
    worksheet.append(title_row_log_reg)
    for row in log_reg_roc_rows:
        worksheet.append(row)
            
    #Precision Recall Curve Data
    pop_xgb = len(thresholds_pr_xgb)
    xgb_pr_df = pd.DataFrame({'Threshold': thresholds_pr_xgb, 'Recall': recall_xgb[:pop_xgb], 'Precision': precision_xgb[:pop_xgb]})
    pop_lr = len(thresholds_pr_log_reg)
    log_reg_pr_df = pd.DataFrame({'Threshold': thresholds_pr_log_reg, 'Recall': recall_log_reg[:pop_lr], 'Precision': precision_log_reg[:pop_lr]})
    log_reg_pr_rows = list(dataframe_to_rows(log_reg_pr_df, index=False))
    xgb_pr_rows = list(dataframe_to_rows(xgb_pr_df, index=False))
    worksheet = workbook.create_sheet(title=f'Precision Recall Curve Data - {target}')
    
    worksheet.append(title_row_xgb)
    for row in xgb_pr_rows:
        worksheet.append(row)
    worksheet.append(title_row_log_reg)
    for row in log_reg_pr_rows:
        worksheet.append(row)

    # Get feature importances for XGBoost
    xgb_feature_importances = xgb.feature_importances_
    feature_importance_df = pd.DataFrame({'Feature': X_train.columns, 'Importance': xgb_feature_importances})
    feature_importance_df = feature_importance_df.sort_values(by='Importance', ascending=False)
#%% Graphical displays for Classification   
    # Confusion Matrices
    conf_log_reg = confusion_matrix(Y_test, log_reg_pred)
    conf_xgb = confusion_matrix(Y_test, xgb_pred)
    
    # Create subplots with four rows, two columns
    fig = make_subplots(
        rows=5, cols=2, vertical_spacing=0.06, 
        subplot_titles=("Feature Importance", "Correlation Matrix", "Confusion Matrix: Logistic Regression", "Confusion Matrix: XGBoost",
                        "ROC Curves", "PR Curves", "Calibration Curves"),
        specs=[[{'colspan':2}, None], [{'colspan':2}, None], [{}, {}], [{}, {}], [{'colspan':2}, None]]
    )

    # Plot ten features with highest importance
    fig.add_trace(go.Bar(x=feature_importance_df.head(10)['Feature'], y=feature_importance_df.head(10)['Importance'], 
                     name='Feature Importance'))
    fig.update_xaxes(title_text='Feature', row=1, col=1)
    fig.update_yaxes(title_text='Importance', row=1, col=1)
    
    # Plot Correlation Matrix
    fig.add_trace(go.Heatmap(z=corr, # Logistic Regression
                             x=corr.columns,
                             y=corr.columns,
                             name='Correlation Matrix',
                             text=corr.values,
                             texttemplate='%{text}',
                             textfont={'size':13},
                             colorbar={'lenmode': 'pixels', 'len':800, 'y':0.711}
                             ), row=2, col=1
                  )

    # Plot Confusion Matrices
    fig.add_trace(go.Heatmap(z=conf_log_reg, # Logistic Regression
                             x=np.unique(Y),
                             y=np.unique(Y),
                             name='CM: Logistic Regression',
                             text=conf_log_reg,
                             texttemplate='%{text}',
                             coloraxis='coloraxis',
                             textfont={'size':25},
                             hoverongaps=False
                             ), row=3, col=1
                  )
    fig.update_xaxes(title_text='Predicted Label', tickvals=Y, ticktext=Y, row=3, col=1)
    fig.update_yaxes(title_text='True Label', tickvals=Y, ticktext=Y, row=3, col=1)

    fig.add_trace(go.Heatmap(z=conf_xgb, # XGBoost
                             x=np.unique(Y),
                             y=np.unique(Y),
                             name='CM: XGBoost',
                             text=conf_xgb,
                             texttemplate='%{text}',
                             textfont={'size':25},
                             coloraxis = 'coloraxis',
                             hoverongaps=False
                             ), row=3, col=2
                  )
    fig.update_xaxes(title_text='Predicted Label', tickvals=Y, row=3, col=2)
    fig.update_yaxes(title_text='True Label', tickvals=Y, row=3, col=2)

    # Plot ROC Curves
    fig.add_trace(go.Scatter(x=fpr_log_reg,
                             y=tpr_log_reg,
                             text=[f'Threshold: {th: .2f}' for th in thresholds_log_reg],
                             hoverinfo='x+y+text',
                             mode='lines',
                             name='ROC: Logistic Regression',
                             marker={'color': 'blue'}
                             ), row=4, col=1
                  ) 
    fig.add_trace(go.Scatter(x=fpr_xgb,
                             y=tpr_xgb,
                             text=[f'Threshold: {th: .2f}' for th in thresholds_xgb],
                             hoverinfo='x+y+text',
                             mode='lines',
                             name='ROC: XGBoost',
                             marker={'color': 'red'}
                             ), row=4,col=1
                  )
    fig.add_shape(
        type='line', line={'dash': 'dash'},
        x0=0, x1=1, y0=0, y1=1,
        opacity=0.5,
        row=4, col=1,
    )
    fig.update_xaxes(title_text='False Positive Rate', row=4, col=1)
    fig.update_yaxes(title_text='True Positive Rate', row=4, col=1)

    # Plot Precision Recall Curves
    fig.add_trace(go.Scatter(x=recall_log_reg,
                             y=precision_log_reg,
                             text=[f'Threshold: {th: .2f}' for th in thresholds_pr_log_reg],
                             hoverinfo='x+y+text',
                             mode='lines',
                             name='Precision-Recall: Logistic Regression',
                             marker={'color': 'blue'}
                             ), row=4, col=2
                  )

    fig.add_trace(go.Scatter(x=recall_xgb,
                             y=precision_xgb,
                             text=[f'Threshold: {th: .2f}' for th in thresholds_pr_xgb],
                             hoverinfo='x+y+text',
                             mode='lines',
                             name='Precision-Recall: XGBoost',
                             marker={'color': 'red'}
                             ), row=4, col=2
                  )
    fig.update_xaxes(title_text='Recall', row=4, col=2)
    fig.update_yaxes(title_text='Precison', row=4, col=2)

    # Plot Calibration Curves
    fig.add_trace(go.Scatter(x=predicted_log_reg_probs,
                             y=true_log_reg_probs,
                             mode='lines',
                             name='Calibration: Logistic Regression',
                             marker={'color': 'blue'}
                             ), row=5, col=1
                  )
    fig.add_trace(go.Scatter(x=predicted_xgb_probs,
                             y=true_xgb_probs,
                             mode='lines',
                             name='Calibration: XGBoost',
                             marker={'color': 'red'}
                             ), row=5, col=1
                  )
    fig.add_shape(
        type='line', line={'dash': 'dash'},
        x0=0, x1=1, y0=0, y1=1,
        opacity=0.5,
        row=5, col=1,
    )
    fig.update_xaxes(title_text='Mean Predicted Probability', row=5, col=1)
    fig.update_yaxes(title_text='True Probability', row=5, col=1)

    # Update layout of the subplots
    fig.update_layout(title={'text': f'{target} & {dataset_filename}', 'x':0.465, 'font_size': 25},
                      height=5400,  # Adjust the height of the combined figure
                      legend={'x':1, 'y':1},  # Hide legends to avoid duplication
                      margin={'l':0, 'r':0, 't': 85, 'b': 0},  # Adjust margins
                      coloraxis={'colorscale': 'Viridis', 'colorbar_lenmode':'pixels', 'colorbar_len':800, 'colorbar_y':0.5},
                      )   

    # Show the plot
    plot(fig, filename=current_location + f'\\{target} Classification.html')
#%% Regression
if modeltype == 'R':
    def Regression(models, QuickSearch):
        scoring_lin = {
            'mean_squared_error': make_scorer(mean_squared_error, greater_is_better=False),
            'r2': make_scorer(r2_score),
            'mean_absolute_percentage_error': make_scorer(mean_absolute_percentage_error, greater_is_better=False),
        }
        
        scoring_xgb = {
            'mean_squared_error': make_scorer(mean_squared_error),
            'r2': make_scorer(r2_score),
            'mean_absolute_percentage_error': make_scorer(mean_absolute_percentage_error),
        }
        results = []
        
        if focus_metric == '':
            refit = 'r2'
        else:
            refit = focus_metric
        
        scaler = StandardScaler()
        X_train_fit = scaler.fit_transform(X_train)
        X_test_fit = scaler.transform(X_test)
        
        for model in models:
            if model == 'linear':
                alg = LinearRegression(n_jobs=-1)
                if QuickSearch == 'Y':
                    # Optimising Linear Regression
                    alg.fit(X_train_fit, Y_train)
                else:
                    lin_reg_space = {
                        'fit_intercept': [True, False],
                        'copy_X': [True, False],
                    }
                    lin_grid_search = GridSearchCV(
                        estimator=alg,
                        param_grid=lin_reg_space,
                        cv=kf,
                        scoring=scoring_lin,
                        refit=refit,
                        n_jobs=-1
                    )
                    lin_grid_search.fit(X_train_fit, Y_train)
                    alg = lin_grid_search.best_estimator_
            else:
                alg = XGBRegressor(n_estimators=1000, learning_rate=0.01, subsample=0.5, n_jobs=-1, random_state=42)
                if QuickSearch != 'Y':
                    # Optimising XGBoost
                    def evaluate_xgb(trial):
                        xgb_space = { # Hyperparameter space for XGBoost
                            'n_estimators': trial.suggest_int('n_estimators', 500, 1500),  # Number of boosting rounds
                            'max_depth': trial.suggest_int('max_depth', 1, 3),  # Maximum depth of a tree
                            'learning_rate': trial.suggest_float('learning_rate', 0.001, 0.01, log=True),  # Step size shrinkage
                            'gamma': trial.suggest_float('gamma', 0.0, 1.0),  # Minimum loss reduction required to make a further partition
                            'colsample_bytree': trial.suggest_float('colsample_bytree', 0.1, 1.0),  # Fraction of features used for fitting the trees
                            'reg_alpha': trial.suggest_float('reg_alpha', 0.0, 100.0),  # L1 regularization term on weights
                            'reg_lambda': trial.suggest_float('reg_lambda', 0.0, 100.0),  # L2 regularization term on weights
                        }
                        xgb_model = XGBRegressor(**xgb_space, subsample=0.5, n_jobs=-1, random_state=42)
                        
                        try:
                            xgb_model.fit(X_train_fit, Y_train)
                        except ValueError:
                            raise optuna.TrialPruned()
                        
                        result = cross_validate(xgb_model, X_train_fit, Y_train, cv=kf, scoring=scoring_xgb, n_jobs=-1)

                        if focus_metric != '':
                            score = result[f'test_{focus_metric}'].mean()
                            return score
                        else: # Make suggestion based on class distributions if there is 
                            r2 = result['test_r2'].mean()
                            return r2
                    if focus_metric == 'mean_squared_error' or focus_metric == 'mean_absolute_percentage_error':
                        direction = 'minimize'
                    else:                    
                        direction = 'maximize'
                    study_xgb = optuna.create_study(direction=direction)
                    study_xgb.optimize(evaluate_xgb, n_trials=200)
                    best_xgb_params = study_xgb.best_params
                    alg = XGBRegressor(**best_xgb_params, random_state=42)
                alg.fit(X_train_fit, Y_train)
            results.append(alg)
            
            pred = alg.predict(X_test_fit)
            results.append(pred)
            residuals = Y_test - pred
            results.append(residuals)
            
        return results

    lin_reg, lin_reg_pred, lin_residuals, xgb, xgb_pred, xgb_residuals = Regression(models, QuickSearch)
    #Prediction data
    prediction = pd.DataFrame({'True Value': Y_test,
                               'XGBoost Prediction': xgb_pred, 'Residual': xgb_residuals,
                               'Linear Regression Prediction': lin_reg_pred, 'Residual ': lin_residuals}
                              )
    pred_rows = list(dataframe_to_rows(prediction, index=False))
    worksheet = workbook.create_sheet(f'Predictive Data - {target}')
    for row in pred_rows:
        worksheet.append(row)
    #Get Feature Importances from XGBoost
    xgb_feature_importances = xgb.feature_importances_
    feature_importance_df = pd.DataFrame({'Feature': X_train.columns, 'Importance': xgb_feature_importances})
    feature_importance_df = feature_importance_df.sort_values(by='Importance', ascending=False)
                    
#%% Graphical displays for Regression 
    # Create subplots with four rows
    fig = make_subplots(rows=5, vertical_spacing=0.06, subplot_titles=("Feature Importance", "Correlation Matrix", "True vs Predicted", "Residuals", "Distribution of Residuals"))
    
    # Plot ten features with highest importance
    fig.add_trace(go.Bar(x=feature_importance_df.head(10)['Feature'], y=feature_importance_df.head(10)['Importance'], 
                         marker={'autocolorscale': False}, name='Feature Importance'))
    fig.update_xaxes(title_text='Feature')
    fig.update_yaxes(title_text='Importance', row=1, col=1)
    
    # Plot Correlation Matrix
    fig.add_trace(go.Heatmap(z=corr,
                             x=corr.columns,
                             y=corr.columns,
                             name='Correlation Matrix',
                             text=corr.values,
                             texttemplate='%{text}',
                             textfont={'size':13},
                             colorbar={'lenmode': 'pixels', 'len':800, 'y':0.711}
                             ), row=2, col=1
                  )

    
    # Plot predicted results against actual results
    fig.add_trace(go.Scatter(x=lin_reg_pred,
                             y=Y_test, mode='markers',
                             name='Linear Regression: True vs Model',
                             marker={'color': 'blue', 'size': 10}
                             ), row=3, col=1
                  )
    fig.add_trace(go.Scatter(x=xgb_pred,
                             y=Y_test,
                             mode='markers',
                             name='XGBoost: True vs Model',
                             marker={'color': 'red', 'size': 10}
                             ), row=3, col=1
                  )

    fig.update_xaxes(title_text=f'Predicted {target}', row=3, col=1)
    fig.update_yaxes(title_text=f'True {target}', row=3, col=1)
    
    # Plot residuals
    fig.add_trace(go.Scatter(x=lin_reg_pred,
                             y=lin_residuals,
                             mode='markers',
                             name='Linear Regression: Residuals ',
                             marker={'color': 'blue', 'size': 10}
                             ), row=4, col=1
                  )
    fig.add_trace(go.Scatter(x=xgb_pred,
                             y=xgb_residuals,
                             mode='markers',
                             name='XGBoost: Residuals',
                             marker={'color': 'red', 'size': 10}
                             ), row=4, col=1
                  )

    fig.update_xaxes(title_text=f'Predicted {target}', row=4, col=1)
    fig.update_yaxes(title_text='Residuals', row=4, col=1)
    
    # Distribution of residuals
    fig.add_trace(go.Histogram(x=lin_residuals,
                               nbinsx=20,
                               name='Linear Regression: Distribution',
                               marker={'color': 'blue'}
                               ), row=5, col=1
                  )
    fig.add_trace(go.Histogram(x=xgb_residuals,
                               nbinsx=20,
                               name='XGBoost: Distribution',
                               marker={'color': 'red', 'opacity': 0.6}
                               ), row=5, col=1
                  )
    fig.update_xaxes(title_text='Residual bins', row=5, col=1)
    fig.update_yaxes(title_text='Frequency', row=5, col=1)
    
    # Update layout of the subplots
    fig.update_layout(title={'text': f'{target} & {dataset_filename}', 'x':0.5, 'font_size': 25},
                      height=5400,  # Adjust the height of the combined figure
                      legend={'x':1, 'y':1},  # Hide legends to avoid duplication
                      margin={'l':0, 'r':0, 't': 85, 'b': 0},  # Adjust margins
                      )
    
    # Show the plot
    plot(fig, filename=current_location + f'\\{target} Regression.html')
#%% Close workbook
#Feature Importance
worksheet = workbook.create_sheet('Feature Importances')
featurerows = list(dataframe_to_rows(feature_importance_df, index=False))
for row in featurerows:
    worksheet.append(row)
workbook.save(current_location + '\\' f'{filename} Modelling.xlsx')
